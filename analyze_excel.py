#!/usr/bin/env python3
import pandas as pd
import openpyxl

def analyze_excel_file(filename):
    print(f"\n=== Analyzing {filename} ===")
    
    wb = openpyxl.load_workbook(filename)
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        try:
            df = pd.read_excel(filename, sheet_name=sheet_name)
            print(f"Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            print("First few rows:")
            print(df.head())
            print("\nData types:")
            print(df.dtypes)
        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")
            try:
                df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
                print(f"Shape (no header): {df.shape}")
                print("First few rows (no header):")
                print(df.head())
            except Exception as e2:
                print(f"Error reading sheet {sheet_name} without header: {e2}")

if __name__ == "__main__":
    analyze_excel_file("kukohakozaki_timetable.xlsx")
    analyze_excel_file("nanakuma_timetable.xlsx")
